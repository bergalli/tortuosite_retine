from __future__ import annotations

import io
import math
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import friedmanchisquare, mannwhitneyu, rankdata, spearmanr, wilcoxon

from tortuosite_score.app.constants import PROJECT_ROOT
from tortuosite_score.vessels_detection.scoring import ScoringConfig, score_run, scoring_method_spec

CLINICAL_METADATA_PATH = PROJECT_ROOT / "demo" / "clinical_metadata.csv"
SHEET_NAMES = ["Resume", "Par_oeil"]
CLASSIFIED_VESSEL_SHEET_NAMES = ["Clean_vessels", "Rejected_vessels"]


@dataclass(frozen=True)
class RunInfo:
    run: str
    patient_id: int | None
    eye: str | None
    group: str


@dataclass(frozen=True)
class VesselNameInfo:
    vessel_name: str
    normalized_name: str
    rank: int | None
    vessel_type: str
    territory: str | None
    is_ambiguous: bool
    issue: str


def generate_clinical_excel(
    run_dirs: list[Path],
    scoring_config: ScoringConfig,
    clinical_metadata_path: Path = CLINICAL_METADATA_PATH,
) -> bytes:
    sheets = build_clinical_analysis_sheets(run_dirs, scoring_config, clinical_metadata_path)
    return _write_excel(sheets, SHEET_NAMES)


def generate_classified_vessels_excel(
    run_dirs: list[Path],
    scoring_config: ScoringConfig,
    clinical_metadata_path: Path = CLINICAL_METADATA_PATH,
) -> bytes:
    data, _quality = build_structured_vessel_data(run_dirs, scoring_config, clinical_metadata_path)
    sheets = build_classified_vessel_sheets(data)
    return _write_excel(sheets, CLASSIFIED_VESSEL_SHEET_NAMES)


def generate_clinical_excel_outputs(
    run_dirs: list[Path],
    scoring_config: ScoringConfig,
    clinical_metadata_path: Path = CLINICAL_METADATA_PATH,
) -> tuple[bytes, bytes]:
    """Generate both workbooks while scoring and classifying every vessel once."""

    data, quality = build_structured_vessel_data(run_dirs, scoring_config, clinical_metadata_path)
    analysis_sheets = _analysis_sheets_from_structured(data, quality)
    classified_sheets = build_classified_vessel_sheets(data)
    return (
        _write_excel(analysis_sheets, SHEET_NAMES),
        _write_excel(classified_sheets, CLASSIFIED_VESSEL_SHEET_NAMES),
    )


def _write_excel(sheets: dict[str, pd.DataFrame], sheet_names: list[str]) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name in sheet_names:
            table = sheets.get(sheet_name, pd.DataFrame())
            table.to_excel(writer, sheet_name=sheet_name, index=False)
            _format_excel_sheet(writer.book[sheet_name], table)
    return buffer.getvalue()


def _format_excel_sheet(worksheet, table: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="17365D")
    summary_fill = PatternFill("solid", fgColor="DCE6F1")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions
    for column_index, column_name in enumerate(table.columns, start=1):
        values = [str(column_name), *[str(value) for value in table[column_name].dropna().head(200)]]
        width = min(42, max(11, max(len(value) for value in values) + 2))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    if "section" in table.columns:
        section_column = table.columns.get_loc("section") + 1
        for row_index in range(2, worksheet.max_row + 1):
            if worksheet.cell(row_index, section_column).value == "resume":
                for cell in worksheet[row_index]:
                    cell.fill = summary_fill
                    cell.font = Font(bold=True)
    if worksheet.title == "Resume":
        worksheet.freeze_panes = "A2"
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_index].height = 72
    if worksheet.title == "Par_oeil":
        positive_fill = PatternFill("solid", fgColor="E2F0D9")
        negative_fill = PatternFill("solid", fgColor="FCE4D6")
        missing_fill = PatternFill("solid", fgColor="E7E6E6")
        headers = {cell.value: cell.column for cell in worksheet[1]}
        for column_name in ["Score_R1", "Score_R2", "Score_R3", "Delta_R3_R1", "Delta_R3_R2"]:
            if column_name in headers:
                for row_index in range(2, worksheet.max_row + 1):
                    worksheet.cell(row_index, headers[column_name]).number_format = "0.00"
        for column_name in ["R3_superieur_R1", "R3_superieur_R2", "R1_R2_R3_disponibles"]:
            if column_name not in headers:
                continue
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row_index, headers[column_name])
                cell.fill = positive_fill if cell.value == "Oui" else negative_fill if cell.value == "Non" else missing_fill
                cell.alignment = Alignment(horizontal="center")


def build_classified_vessel_sheets(data: pd.DataFrame) -> dict[str, pd.DataFrame]:
    long_format = _classified_long_format(data)
    if long_format.empty:
        return {
            "Clean_vessels": long_format.copy(),
            "Rejected_vessels": long_format.copy(),
        }
    eligible = long_format["classified_export_eligible"].fillna(False).astype(bool)
    return {
        "Clean_vessels": long_format[eligible].reset_index(drop=True),
        "Rejected_vessels": long_format[~eligible].reset_index(drop=True),
    }


def _classified_long_format(data: pd.DataFrame) -> pd.DataFrame:
    if data.empty:
        return pd.DataFrame(columns=_raw_export_columns())
    table = data.copy()
    rank_label = table["rang"].apply(lambda value: f"R{int(value)}" if pd.notna(value) else "sans_rang")
    territory_label = table["territoire"].fillna("sans_territoire").replace("", "sans_territoire")
    table["classe_vaisseau"] = table["type_vaisseau"].fillna("unknown") + "_" + rank_label + "_" + territory_label
    table = table.sort_values(
        ["patient_id", "groupe", "oeil", "run", "type_vaisseau", "rang", "territoire", "vessel_name_normalise"],
        na_position="last",
    ).reset_index(drop=True)
    table["numero_dans_classe"] = table.groupby(["run", "classe_vaisseau"], dropna=False).cumcount() + 1
    table = table.rename(
        columns={
            "groupe": "group",
            "oeil": "eye",
            "numero_dans_classe": "class_number",
            "rang": "rank",
            "type_vaisseau": "vessel_type",
            "territoire": "territory",
            "score_brut": "score_raw",
            "score": "score_normalized",
            "longueur_brute_px": "length_raw_px",
            "longueur": "length_normalized_px",
            "facteur_normalisation": "normalization_factor",
            "diametre_fond_oeil_px": "fundus_diameter_px",
            "diametre_reference_px": "reference_diameter_px",
            "eligible_analyse": "analysis_eligible",
            "raison_exclusion_analyse": "analysis_exclusion_reason",
            "eligible_export_classifie": "classified_export_eligible",
            "raison_exclusion_export_classifie": "classified_export_exclusion_reason",
            "score_local_bump_v1": "local_bump_v1_score",
            "score_local_bump_v2": "local_bump_v2_score",
            "composante_oscillation_v2": "local_bump_v2_oscillation_component",
            "composante_angularite_v2": "local_bump_v2_angularity_component",
            "nombre_lobes_persistants": "persistent_lobe_count",
            "nombre_oscillations_persistantes": "persistent_oscillation_count",
            "tour_maximal_extremites": "endpoint_max_turn",
            "segments_modele": "model_segment_count",
            "segments_manuels": "manual_segment_count",
        }
    )
    table["group"] = table["group"].replace({"temoin": "control"})
    table["eye"] = table["eye"].replace({"OD": "right", "OG": "left"})
    table["vessel_type"] = table["vessel_type"].replace({"artere": "artery", "veine": "vein"})
    table["territory"] = table["territory"].replace({"inf": "inferior", "sup": "superior"})
    return table[_raw_export_columns()]


def _raw_export_columns() -> list[str]:
    return [
        # Subject and eye characteristics.
        "patient_id",
        "group",
        "eye",
        # Original identifiers.
        "run",
        "vessel_name",
        # Derived clinical classification.
        "class_number",
        "vessel_type",
        "rank",
        "territory",
        # Measured and normalization values.
        "length_raw_px",
        "length_normalized_px",
        "fundus_diameter_px",
        "reference_diameter_px",
        "normalization_factor",
        "analysis_eligible",
        "analysis_exclusion_reason",
        "classified_export_eligible",
        "classified_export_exclusion_reason",
        # Final score values.
        "score_raw",
        "score_normalized",
        # Auditable local-bump v1/v2 components and geometry source.
        "local_bump_v1_score",
        "local_bump_v2_score",
        "local_bump_v2_oscillation_component",
        "local_bump_v2_angularity_component",
        "persistent_lobe_count",
        "persistent_oscillation_count",
        "endpoint_max_turn",
        "model_segment_count",
        "manual_segment_count",
    ]


def build_clinical_analysis_sheets(
    run_dirs: list[Path],
    scoring_config: ScoringConfig,
    clinical_metadata_path: Path = CLINICAL_METADATA_PATH,
) -> dict[str, pd.DataFrame]:
    data, quality = build_structured_vessel_data(run_dirs, scoring_config, clinical_metadata_path)
    return _analysis_sheets_from_structured(data, quality)


def _analysis_sheets_from_structured(data: pd.DataFrame, _quality: pd.DataFrame) -> dict[str, pd.DataFrame]:
    rank_scores = _rank_eye_scores(data)
    return _readable_same_eye_rank_sheets(rank_scores)


def build_structured_vessel_data(
    run_dirs: list[Path],
    scoring_config: ScoringConfig,
    clinical_metadata_path: Path = CLINICAL_METADATA_PATH,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    method = scoring_method_spec(scoring_config.method_id)
    scale = float(method.eye_score_scale)
    clinical = _read_clinical_metadata(clinical_metadata_path)
    clinical_by_patient = clinical.set_index("patient_id") if not clinical.empty else pd.DataFrame()
    rows: list[dict[str, object]] = []
    quality_rows: list[dict[str, object]] = []
    seen_patient_eyes: set[tuple[int, str]] = set()
    seen_control_eyes: set[tuple[int, str]] = set()
    run_infos: list[RunInfo] = []

    for run_dir in sorted(run_dirs, key=lambda path: path.name):
        run_info = parse_run_name(run_dir.name)
        run_infos.append(run_info)
        if run_info.patient_id is not None and run_info.eye is not None:
            if run_info.group == "patient":
                seen_patient_eyes.add((run_info.patient_id, run_info.eye))
            elif run_info.group == "temoin":
                seen_control_eyes.add((run_info.patient_id, run_info.eye))
        if not (run_dir / "manual_review_state.json").exists():
            quality_rows.append(_quality_row(run_info, "", "run_sans_etat_manuel", "manual_review_state.json manquant"))
        try:
            _summary, vessel_scores = score_run(run_dir, scoring_config)
        except (FileNotFoundError, ValueError, OSError, KeyError) as exc:
            quality_rows.append(_quality_row(run_info, "", "run_non_exploitable", str(exc)))
            continue
        if vessel_scores.empty:
            quality_rows.append(_quality_row(run_info, "", "aucun_vaisseau_score", "Aucun vaisseau sauvegarde score"))
            continue
        for _, score_row in vessel_scores.iterrows():
            vessel_name = str(score_row.get("vessel_name", ""))
            vessel_info = parse_vessel_name(vessel_name, score_row.get("category"))
            clinical_values = _clinical_values(clinical_by_patient, run_info.patient_id)
            score = _finite_float(score_row.get("primary_score"))
            raw_score = _finite_float(score_row.get("raw_primary_score", score_row.get("primary_score")))
            length = _finite_float(score_row.get("vessel_length"))
            length_eligible = bool(score_row.get("meets_length_threshold", score_row.get("eligible", True)))
            scoring_eligible = bool(score_row.get("eligible", False))
            analysis_eligible, exclusion_reason = analysis_vessel_eligibility(
                vessel_name,
                score_row.get("category"),
                length,
            )
            valid_clinical_name = _is_ranked_artery(vessel_info)
            usable_geometry = math.isfinite(length) and length > 0
            classified_export_eligible, classified_export_exclusion_reason = (
                classified_vessel_export_eligibility(
                    vessel_info,
                    analysis_eligible=analysis_eligible,
                    scoring_eligible=scoring_eligible,
                    score=score,
                    vessel_length=length,
                )
            )
            rows.append(
                {
                    "patient_id": run_info.patient_id,
                    "groupe": run_info.group,
                    "oeil": run_info.eye,
                    "run": run_info.run,
                    "vessel_name": vessel_name,
                    "vessel_name_normalise": vessel_info.normalized_name,
                    "rang": vessel_info.rank,
                    "type_vaisseau": vessel_info.vessel_type,
                    "territoire": vessel_info.territory,
                    "score_brut": raw_score * scale if math.isfinite(raw_score) else math.nan,
                    "score": score * scale if math.isfinite(score) else math.nan,
                    "longueur": length,
                    "longueur_brute_px": _finite_float(score_row.get("raw_vessel_length")),
                    "facteur_normalisation": _finite_float(score_row.get("coordinate_scale")),
                    "diametre_fond_oeil_px": _finite_float(score_row.get("fundus_diameter_px")),
                    "diametre_reference_px": _finite_float(score_row.get("reference_fundus_diameter_px")),
                    "eligible": length_eligible,
                    "eligible_filtre_longueur": length_eligible,
                    "nom_clinique_valide": valid_clinical_name,
                    "eligible_analyse": analysis_eligible,
                    "raison_exclusion_analyse": exclusion_reason,
                    "eligible_export_classifie": classified_export_eligible,
                    "raison_exclusion_export_classifie": classified_export_exclusion_reason,
                    "methode_id": method.method_id,
                    "methode": score_row.get("scoring_method_label"),
                    "score_local_bump_v1": _finite_float(score_row.get("local_bump_score")) * 1000.0,
                    "score_local_bump_v2": _finite_float(score_row.get("local_bump_v2_score")),
                    "composante_oscillation_v2": _finite_float(
                        score_row.get("local_bump_v2_oscillation_component")
                    ),
                    "composante_angularite_v2": _finite_float(
                        score_row.get("local_bump_v2_angularity_component")
                    ),
                    "nombre_lobes_persistants": _finite_float(score_row.get("persistent_lobe_count")),
                    "nombre_oscillations_persistantes": _finite_float(
                        score_row.get("persistent_oscillation_count")
                    ),
                    "tour_maximal_extremites": _finite_float(score_row.get("endpoint_max_turn")),
                    "segments_modele": _finite_float(score_row.get("model_segment_count")),
                    "segments_manuels": _finite_float(score_row.get("manual_segment_count")),
                    "tortuosity_density_score": _finite_float(score_row.get("tortuosity_density_score")),
                    "sous_segments_courbure_constante": _finite_float(
                        score_row.get("constant_curvature_segment_count")
                    ),
                    "nombre_inflexions": _finite_float(score_row.get("inflection_count")),
                    "exces_arc_corde_tortuosity_density": _finite_float(
                        score_row.get("tortuosity_density_excess")
                    ),
                    "geometrie_tortuosity_density_valide": bool(
                        score_row.get("tortuosity_density_valid", False)
                    ),
                    "poids_global": scoring_config.local_bump_settings.global_weight,
                    "poids_queue_haute": scoring_config.local_bump_settings.tail_weight,
                    "fraction_queue_haute": scoring_config.local_bump_settings.tail_length_fraction,
                    "age": clinical_values.get("age", math.nan),
                    "severite": clinical_values.get("severite", math.nan),
                }
            )
            if vessel_info.issue:
                quality_rows.append(_quality_row(run_info, vessel_name, vessel_info.issue, "Nom de vaisseau a verifier"))
            if not length_eligible and valid_clinical_name and usable_geometry:
                quality_rows.append(
                    _quality_row(
                        run_info,
                        vessel_name,
                        "vaisseau_court_conserve_nom_valide",
                        f"Longueur normalisee {length:.1f}; conserve car le nom permet une classification clinique",
                    )
                )
            elif not usable_geometry:
                quality_rows.append(
                    _quality_row(
                        run_info,
                        vessel_name,
                        "vaisseau_exclu_longueur_nulle",
                        "Longueur nulle ou non calculable",
                    )
                )
            elif not analysis_eligible:
                quality_rows.append(
                    _quality_row(
                        run_info,
                        vessel_name,
                        "vaisseau_exclu_nom_non_classe",
                        "Nom non classe comme artere de rang 1, 2 ou 3",
                    )
                )

    _append_cohort_quality_rows(quality_rows, run_infos, seen_patient_eyes, seen_control_eyes, clinical)
    data = pd.DataFrame(rows, columns=_structured_columns())
    _append_resolution_quality_rows(quality_rows, data)
    quality = pd.DataFrame(quality_rows, columns=["patient_id", "groupe", "oeil", "run", "vessel_name", "probleme", "detail"])
    return data, quality


def parse_run_name(run_name: str) -> RunInfo:
    patient_match = re.fullmatch(r"(\d+)_(OD|OG)", run_name, flags=re.IGNORECASE)
    if patient_match:
        return RunInfo(run=run_name, patient_id=int(patient_match.group(1)), eye=patient_match.group(2).upper(), group="patient")
    control_match = re.fullmatch(r"(OD|OG)_de_(\d+)", run_name, flags=re.IGNORECASE)
    if control_match:
        return RunInfo(run=run_name, patient_id=int(control_match.group(2)), eye=control_match.group(1).upper(), group="temoin")
    return RunInfo(run=run_name, patient_id=None, eye=None, group="unknown")


def parse_vessel_name(vessel_name: str, category: object | None = None) -> VesselNameInfo:
    normalized = _normalize_name(vessel_name)
    rank = _artery_rank(vessel_name, normalized)
    looks_vein = bool(re.search(r"\b(v|vein|veine|vine)\b|veine|vein|vine", normalized))
    looks_artery = rank is not None or bool(re.search(r"\b(a|artere|artery)\b|artere|artery|\bati\b|\bats\b", normalized))
    if not looks_vein and not looks_artery:
        category_text = _normalize_name(str(category or ""))
        if "veine" in category_text:
            looks_vein = True
        if "artere" in category_text:
            looks_artery = True
    vessel_type = "unknown"
    if looks_vein and not looks_artery:
        vessel_type = "veine"
    elif looks_artery and not looks_vein:
        vessel_type = "artere"
    elif looks_artery and looks_vein:
        vessel_type = "unknown"
    territory = None
    has_inf = bool(re.search(r"\binf", normalized))
    has_sup = bool(re.search(r"\bsup", normalized))
    if has_inf and not has_sup:
        territory = "inf"
    elif has_sup and not has_inf:
        territory = "sup"
    is_ambiguous = looks_artery and looks_vein or (has_inf and has_sup)
    issue = ""
    if is_ambiguous:
        issue = "nom_ambigu"
    elif vessel_type == "unknown":
        issue = "type_non_classe"
    elif vessel_type == "artere" and rank is None:
        issue = "rang_arteriel_non_classe"
    elif vessel_type == "veine" and rank is not None:
        issue = "veine_avec_rang_detecte"
    return VesselNameInfo(vessel_name, normalized, rank if vessel_type == "artere" else None, vessel_type, territory, is_ambiguous, issue)


def _is_ranked_artery(vessel: VesselNameInfo) -> bool:
    if vessel.is_ambiguous:
        return False
    return vessel.vessel_type == "artere" and vessel.rank in {1, 2, 3}


def analysis_vessel_eligibility(
    vessel_name: str,
    category: object | None,
    vessel_length: object,
) -> tuple[bool, str]:
    """Return the exact eligibility and exclusion reason used by Clean_vessels."""

    length = _finite_float(vessel_length)
    if not math.isfinite(length) or length <= 0:
        return False, "zero_length"
    if not _is_ranked_artery(parse_vessel_name(vessel_name, category)):
        return False, "unclassified_name"
    return True, ""


def classified_vessel_export_eligibility(
    vessel_info: VesselNameInfo,
    *,
    analysis_eligible: bool,
    scoring_eligible: bool,
    score: object,
    vessel_length: object,
) -> tuple[bool, str]:
    """Return eligibility for the classified raw workbook without changing rank analysis."""

    if analysis_eligible:
        return True, ""
    length = _finite_float(vessel_length)
    if not math.isfinite(length) or length <= 0:
        return False, "zero_length"
    if vessel_info.vessel_type != "veine" or vessel_info.is_ambiguous:
        return False, "unclassified_type"
    if not math.isfinite(_finite_float(score)):
        return False, "invalid_score"
    if not scoring_eligible:
        return False, "scoring_ineligible"
    return True, ""


def _artery_rank(original_name: str, normalized: str) -> int | None:
    original = original_name.lower()
    patterns = [
        r"(^|\b)([123])\s*°?\s*a\b",
        r"\bartere[_\s-]*([123])\b",
        r"\b([123])\s*(?:er|eme|e)\b.*\b(?:artere|ati|ats)\b",
    ]
    for pattern in patterns:
        for text in [original, normalized]:
            match = re.search(pattern, text, flags=re.IGNORECASE)
            if match:
                return int(match.group(match.lastindex))
    return None


def _readable_same_eye_rank_sheets(rank_scores: pd.DataFrame) -> dict[str, pd.DataFrame]:
    patient_scores = rank_scores[rank_scores["groupe"] == "patient"].copy() if not rank_scores.empty else pd.DataFrame()
    eye_table = _rank_pivot(patient_scores, ["patient_id", "oeil", "run"])
    if eye_table.empty:
        return {
            "Resume": pd.DataFrame(
                [
                    {
                        "Question": "Les arteres de rang 3 sont-elles plus tortueuses que celles de rang 1 et 2 dans un meme oeil ?",
                        "Conclusion": "Donnees insuffisantes",
                    }
                ]
            ),
            "Par_oeil": pd.DataFrame(),
        }

    eye_table["delta_R3_R1"] = eye_table["score_rang_3"] - eye_table["score_rang_1"]
    eye_table["delta_R3_R2"] = eye_table["score_rang_3"] - eye_table["score_rang_2"]
    eye_table["R3_superieur_R1"] = eye_table["delta_R3_R1"].apply(_yes_no_missing)
    eye_table["R3_superieur_R2"] = eye_table["delta_R3_R2"].apply(_yes_no_missing)
    eye_table["donnees_completes_R1_R2_R3"] = eye_table[["score_rang_1", "score_rang_2", "score_rang_3"]].notna().all(axis=1).map({True: "Oui", False: "Non"})

    summary_rows: list[dict[str, object]] = []
    comparison_specs = [
        (1, "R3 versus R1", "score_rang_1", "delta_R3_R1"),
        (2, "R3 versus R2", "score_rang_2", "delta_R3_R2"),
    ]
    paired_results: list[dict[str, object]] = []
    for rank, label, comparator_column, delta_column in comparison_specs:
        eye_pairs = eye_table.dropna(subset=["score_rang_3", comparator_column]).copy()
        patient_pairs = (
            eye_pairs.groupby("patient_id", as_index=False)[["score_rang_3", comparator_column]]
            .mean()
        )
        result = _paired_summary(
            patient_pairs["score_rang_3"],
            patient_pairs[comparator_column],
            label,
            alternative="greater",
        )
        result["_rank"] = rank
        result["_eye_pairs"] = eye_pairs
        result["_patient_pairs"] = patient_pairs
        result["_delta_column"] = delta_column
        paired_results.append(result)

    _holm_adjust_summary_rows(paired_results, "p_value_wilcoxon", "p_value_wilcoxon_holm")
    _holm_adjust_summary_rows(paired_results, "p_value_directionnelle", "p_value_directionnelle_holm")
    for result in paired_results:
        rank = int(result.pop("_rank"))
        eye_pairs = result.pop("_eye_pairs")
        patient_pairs = result.pop("_patient_pairs")
        delta_column = str(result.pop("_delta_column"))
        eye_positive = int((eye_pairs[delta_column] > 0).sum())
        patient_differences = patient_pairs["score_rang_3"] - patient_pairs[f"score_rang_{rank}"]
        patient_positive = int((patient_differences > 0).sum())
        adjusted_p = _finite_float(result.get("p_value_wilcoxon_holm"))
        directional_adjusted_p = _finite_float(result.get("p_value_directionnelle_holm"))
        delta = _finite_float(result.get("delta_moyen"))
        summary_rows.append(
            {
                "Question": "Le rang 3 est-il plus tortueux dans un meme oeil ?",
                "Comparaison": f"R3 vs R{rank}",
                "Population": "Patients; differences calculees dans chaque oeil puis moyennees par patient pour le test",
                "N_patients": int(len(patient_pairs)),
                "N_yeux": int(len(eye_pairs)),
                "Score_moyen_R3_par_patient": float(patient_pairs["score_rang_3"].mean()),
                "Score_moyen_comparateur_par_patient": float(patient_pairs[f"score_rang_{rank}"].mean()),
                "Delta_moyen_par_patient": delta,
                "Delta_median_patient": _finite_float(result.get("delta_median")),
                "Yeux_avec_R3_superieur": _count_percent(eye_positive, len(eye_pairs)),
                "Patients_avec_R3_superieur": _count_percent(patient_positive, len(patient_pairs)),
                "IC95_delta_moyen_bas": _finite_float(result.get("ic95_delta_moyen_bas")),
                "IC95_delta_moyen_haut": _finite_float(result.get("ic95_delta_moyen_haut")),
                "p_Wilcoxon_bilaterale": _finite_float(result.get("p_value_wilcoxon")),
                "p_bilaterale_corrigee_Holm": adjusted_p,
                "p_directionnelle_R3_superieur": _finite_float(result.get("p_value_directionnelle")),
                "p_directionnelle_corrigee_Holm": directional_adjusted_p,
                "Conclusion": _rank_comparison_conclusion(adjusted_p, directional_adjusted_p, delta, eye_positive, len(eye_pairs), rank),
            }
        )

    detail = eye_table.rename(
        columns={
            "patient_id": "Patient",
            "oeil": "Oeil",
            "run": "Run",
            "score_rang_1": "Score_R1",
            "score_rang_2": "Score_R2",
            "score_rang_3": "Score_R3",
            "n_rang_1": "Nb_vaisseaux_R1",
            "n_rang_2": "Nb_vaisseaux_R2",
            "n_rang_3": "Nb_vaisseaux_R3",
            "delta_R3_R1": "Delta_R3_R1",
            "delta_R3_R2": "Delta_R3_R2",
            "donnees_completes_R1_R2_R3": "R1_R2_R3_disponibles",
        }
    )
    detail_columns = [
        "Patient",
        "Oeil",
        "Run",
        "Score_R1",
        "Score_R2",
        "Score_R3",
        "Delta_R3_R1",
        "R3_superieur_R1",
        "Delta_R3_R2",
        "R3_superieur_R2",
        "Nb_vaisseaux_R1",
        "Nb_vaisseaux_R2",
        "Nb_vaisseaux_R3",
        "R1_R2_R3_disponibles",
    ]
    for count_column in ["Nb_vaisseaux_R1", "Nb_vaisseaux_R2", "Nb_vaisseaux_R3"]:
        detail[count_column] = pd.to_numeric(detail[count_column], errors="coerce").astype("Int64")
    return {"Resume": pd.DataFrame(summary_rows), "Par_oeil": detail[detail_columns]}


def _yes_no_missing(value: object) -> str:
    numeric = _finite_float(value)
    if not math.isfinite(numeric):
        return "NA"
    return "Oui" if numeric > 0 else "Non"


def _count_percent(count: int, total: int) -> str:
    return f"{count}/{total} ({count / total:.1%})" if total else "0/0 (NA)"


def _rank_comparison_conclusion(
    adjusted_p: float,
    directional_adjusted_p: float,
    delta: float,
    positive_eyes: int,
    total_eyes: int,
    rank: int,
) -> str:
    proportion = positive_eyes / total_eyes if total_eyes else math.nan
    if math.isfinite(adjusted_p) and adjusted_p < 0.05 and delta > 0:
        return f"R3 est statistiquement plus eleve que R{rank}, mais pas dans tous les yeux ({proportion:.0%})."
    if math.isfinite(directional_adjusted_p) and directional_adjusted_p < 0.05 and delta > 0:
        return f"Signal directionnel en faveur de R3 > R{rank}, sans preuve bilaterale; R3 n'est pas superieur dans tous les yeux ({proportion:.0%})."
    return f"R3 n'est pas demontre plus eleve que R{rank} et n'est pas superieur dans tous les yeux ({proportion:.0%})."


def _same_eye_rank_sheet(rank_scores: pd.DataFrame) -> pd.DataFrame:
    patient_scores = rank_scores[rank_scores["groupe"] == "patient"].copy() if not rank_scores.empty else pd.DataFrame()
    eye_pivot = _rank_pivot(patient_scores, ["patient_id", "oeil", "run"])
    if eye_pivot.empty:
        return _with_summary(pd.DataFrame(), "donnees insuffisantes")
    eye_pivot["rang_3_plus_eleve"] = eye_pivot.apply(
        lambda row: _rank3_is_highest(row.get("score_rang_1"), row.get("score_rang_2"), row.get("score_rang_3")),
        axis=1,
    )
    eye_pivot["delta_R3_R1"] = eye_pivot["score_rang_3"] - eye_pivot["score_rang_1"]
    eye_pivot["delta_R3_R2"] = eye_pivot["score_rang_3"] - eye_pivot["score_rang_2"]
    eye_pivot["section"] = "detail_oeil"

    complete_eyes = eye_pivot.dropna(subset=["score_rang_1", "score_rang_2", "score_rang_3"])
    patient_pivot = (
        complete_eyes.groupby("patient_id", as_index=False)[["score_rang_1", "score_rang_2", "score_rang_3"]]
        .mean()
    )
    eye_counts = complete_eyes.groupby("patient_id").size().rename("n_yeux_complets").reset_index()
    patient_pivot = patient_pivot.merge(eye_counts, on="patient_id", how="left")
    patient_pivot["delta_R3_R1"] = patient_pivot["score_rang_3"] - patient_pivot["score_rang_1"]
    patient_pivot["delta_R3_R2"] = patient_pivot["score_rang_3"] - patient_pivot["score_rang_2"]
    patient_pivot["rang_3_plus_eleve"] = patient_pivot.apply(
        lambda row: _rank3_is_highest(row.get("score_rang_1"), row.get("score_rang_2"), row.get("score_rang_3")),
        axis=1,
    )
    patient_pivot["section"] = "detail_patient"
    summary = [
        _paired_summary(patient_pivot["score_rang_3"], patient_pivot["score_rang_1"], "R3_vs_R1", alternative="greater"),
        _paired_summary(patient_pivot["score_rang_3"], patient_pivot["score_rang_2"], "R3_vs_R2", alternative="greater"),
    ]
    complete = patient_pivot.dropna(subset=["score_rang_1", "score_rang_2", "score_rang_3"])
    if len(complete) >= 3:
        friedman = friedmanchisquare(complete["score_rang_1"], complete["score_rang_2"], complete["score_rang_3"])
        summary.insert(
            0,
            {
                "section": "resume",
                "comparaison": "R1_vs_R2_vs_R3_omnibus_patient",
                "n_paires": int(len(complete)),
                "statistique_friedman": float(friedman.statistic),
                "p_value": float(friedman.pvalue),
                "conclusion": _omnibus_conclusion(float(friedman.pvalue)),
            },
        )
    _holm_adjust_summary_rows(summary, "p_value_wilcoxon", "p_value_wilcoxon_holm")
    _holm_adjust_summary_rows(summary, "p_value_directionnelle", "p_value_directionnelle_holm")
    _refresh_comparison_conclusions(summary, "p_value_wilcoxon_holm")
    return _combine_sections(summary, patient_pivot, eye_pivot)


def _patient_vs_control_sheet(rank_scores: pd.DataFrame) -> pd.DataFrame:
    if rank_scores.empty:
        return _with_summary(pd.DataFrame(), "donnees insuffisantes")
    patients = rank_scores[rank_scores["groupe"] == "patient"]
    controls = rank_scores[rank_scores["groupe"] == "temoin"]
    merged = patients.merge(
        controls,
        on=["patient_id", "oeil", "rang"],
        suffixes=("_patient", "_temoin"),
    )
    if merged.empty:
        return _with_summary(pd.DataFrame(), "aucun appariement patient-temoin")
    table = merged[
        [
            "patient_id",
            "oeil",
            "rang",
            "run_patient",
            "run_temoin",
            "score_principal_patient",
            "score_principal_temoin",
            "score_pondere_patient",
            "score_pondere_temoin",
            "n_vaisseaux_patient",
            "n_vaisseaux_temoin",
        ]
    ].copy()
    table["delta_patient_temoin"] = table["score_principal_patient"] - table["score_principal_temoin"]
    summary = [
        _paired_summary(
            subset["score_principal_patient"],
            subset["score_principal_temoin"],
            f"rang_{rank}_patient_vs_temoin",
            alternative="greater",
        )
        for rank, subset in table.groupby("rang")
    ]
    _holm_adjust_summary_rows(summary, "p_value_wilcoxon", "p_value_wilcoxon_holm")
    _holm_adjust_summary_rows(summary, "p_value_directionnelle", "p_value_directionnelle_holm")
    _refresh_comparison_conclusions(summary, "p_value_wilcoxon_holm")
    return _append_summary_rows(table, summary)


def _od_vs_og_sheet(rank_scores: pd.DataFrame) -> pd.DataFrame:
    if rank_scores.empty:
        return _with_summary(pd.DataFrame(), "donnees insuffisantes")
    patient_scores = rank_scores[rank_scores["groupe"] == "patient"]
    od = patient_scores[patient_scores["oeil"] == "OD"]
    og = patient_scores[patient_scores["oeil"] == "OG"]
    merged = od.merge(og, on=["patient_id", "rang"], suffixes=("_OD", "_OG"))
    if merged.empty:
        return _with_summary(pd.DataFrame(), "aucun appariement OD-OG")
    table = merged[
        [
            "patient_id",
            "rang",
            "run_OD",
            "run_OG",
            "score_principal_OD",
            "score_principal_OG",
            "score_pondere_OD",
            "score_pondere_OG",
            "n_vaisseaux_OD",
            "n_vaisseaux_OG",
        ]
    ].copy()
    table["delta_OD_OG"] = table["score_principal_OD"] - table["score_principal_OG"]
    summary: list[dict[str, object]] = []
    for rank, subset in table.groupby("rang"):
        summary.append(_paired_summary(subset["score_principal_OD"], subset["score_principal_OG"], f"rang_{rank}_difference_OD_vs_OG"))
        pairs = subset[["score_principal_OD", "score_principal_OG"]].dropna()
        if len(pairs) >= 3:
            association = spearmanr(pairs["score_principal_OD"], pairs["score_principal_OG"])
            rho = float(association.statistic)
            pvalue = float(association.pvalue)
        else:
            rho = math.nan
            pvalue = math.nan
        summary.append(
            {
                "section": "resume",
                "comparaison": f"rang_{rank}_association_OD_OG",
                "n_paires": int(len(pairs)),
                "rho_spearman": rho,
                "p_value_spearman": pvalue,
                "conclusion": _bilateral_association_conclusion(pvalue, rho),
            }
        )
    _holm_adjust_summary_rows(summary, "p_value_wilcoxon", "p_value_wilcoxon_holm")
    _holm_adjust_summary_rows(summary, "p_value_spearman", "p_value_spearman_holm")
    _refresh_comparison_conclusions(summary, "p_value_wilcoxon_holm")
    for row in summary:
        if "p_value_spearman_holm" in row:
            row["conclusion"] = _bilateral_association_conclusion(
                _finite_float(row["p_value_spearman_holm"]),
                _finite_float(row.get("rho_spearman")),
            )
    return _append_summary_rows(table, summary)


def _arteries_vs_veins_sheet(data: pd.DataFrame) -> pd.DataFrame:
    if data.empty:
        return _with_summary(pd.DataFrame(), "donnees insuffisantes")
    eligibility_column = "eligible_analyse" if "eligible_analyse" in data else "eligible"
    eligible = data[eligibility_column].fillna(False).astype(bool) if eligibility_column in data else pd.Series(True, index=data.index)
    patient_data = data[(data["groupe"] == "patient") & eligible].copy()
    rows: list[dict[str, object]] = []
    for (patient_id, eye, run), subset in patient_data.groupby(["patient_id", "oeil", "run"], dropna=False):
        veins = subset[subset["type_vaisseau"] == "veine"]
        vein_score = _primary_group_score(veins)
        for rank in [1, 2, 3]:
            arteries = subset[(subset["type_vaisseau"] == "artere") & (subset["rang"] == rank)]
            artery_score = _primary_group_score(arteries)
            rows.append(
                {
                    "patient_id": patient_id,
                    "oeil": eye,
                    "run": run,
                    "rang_arteriel": rank,
                    "score_arteres": artery_score,
                    "score_veines": vein_score,
                    "delta_arteres_veines": artery_score - vein_score if math.isfinite(artery_score) and math.isfinite(vein_score) else math.nan,
                    "n_arteres": int(len(arteries)),
                    "n_veines": int(len(veins)),
                    "focus_rang_3": rank == 3,
                }
            )
    table = pd.DataFrame(rows)
    summary: list[dict[str, object]] = []
    for rank, subset in table.groupby("rang_arteriel"):
        summary.append(
            _paired_summary(
                subset["score_arteres"],
                subset["score_veines"],
                f"rang_{rank}_arteres_vs_veines",
                alternative="greater",
            )
        )
    rank3 = table[table["rang_arteriel"] == 3]
    artery_values = _finite_series(rank3["score_arteres"]) if not rank3.empty else pd.Series(dtype=float)
    vein_values = _finite_series(rank3["score_veines"]) if not rank3.empty else pd.Series(dtype=float)
    summary.append(_mannwhitney_summary(artery_values, vein_values, "rang_3_arteres_vs_veines_nonapparie"))
    _holm_adjust_summary_rows(summary, "p_value_wilcoxon", "p_value_wilcoxon_holm")
    _holm_adjust_summary_rows(summary, "p_value_directionnelle", "p_value_directionnelle_holm")
    _refresh_comparison_conclusions(summary, "p_value_wilcoxon_holm")
    return _append_summary_rows(table, summary)


def _correlation_sheet(rank_scores: pd.DataFrame, data: pd.DataFrame, variable: str) -> pd.DataFrame:
    max_scores = _patient_max_rank_scores(rank_scores)
    if max_scores.empty:
        return _with_summary(pd.DataFrame(), "donnees insuffisantes")
    clinical = data[data["groupe"] == "patient"][["patient_id", variable]].drop_duplicates("patient_id") if not data.empty else pd.DataFrame()
    table = max_scores.merge(clinical, on="patient_id", how="left")
    values = table.dropna(subset=["score_max_rangs_1_2_3", variable])
    if len(values) >= 3:
        stat = spearmanr(values[variable], values["score_max_rangs_1_2_3"], nan_policy="omit")
        rho = float(stat.statistic)
        pvalue = float(stat.pvalue)
        conclusion = _correlation_conclusion(pvalue, rho, variable)
    else:
        rho = math.nan
        pvalue = math.nan
        conclusion = "donnees insuffisantes"
    summary = pd.DataFrame(
        [
            {
                "section": "resume",
                "variable": variable,
                "n_patients_inclus": int(len(values)),
                "rho_spearman": rho,
                "p_value": pvalue,
                "conclusion": conclusion,
            }
        ]
    )
    return pd.concat([table, summary], ignore_index=True, sort=False)


def _rank_eye_scores(data: pd.DataFrame) -> pd.DataFrame:
    if data.empty:
        return pd.DataFrame(columns=["patient_id", "groupe", "oeil", "run", "rang", "score_principal", "score_pondere", "score_queue_haute", "score_median", "n_vaisseaux"])
    eligibility_column = "eligible_analyse" if "eligible_analyse" in data else "eligible"
    eligible = data[eligibility_column].fillna(False).astype(bool) if eligibility_column in data else pd.Series(True, index=data.index)
    subset = data[(data["type_vaisseau"] == "artere") & (data["rang"].isin([1, 2, 3])) & eligible].copy()
    rows: list[dict[str, object]] = []
    for (patient_id, group, eye, run, rank), group_df in subset.groupby(["patient_id", "groupe", "oeil", "run", "rang"], dropna=False):
        weighted_score = _weighted_group_score(group_df)
        tail_score = _tail_group_score(group_df)
        method_id = str(group_df["methode_id"].iloc[0]) if "methode_id" in group_df else ""
        if method_id == "local_bump" and math.isfinite(tail_score):
            global_weight = _finite_float(group_df["poids_global"].iloc[0])
            tail_weight = _finite_float(group_df["poids_queue_haute"].iloc[0])
            primary_score = global_weight * weighted_score + tail_weight * tail_score
        else:
            primary_score = weighted_score
        rows.append(
            {
                "patient_id": patient_id,
                "groupe": group,
                "oeil": eye,
                "run": run,
                "rang": int(rank),
                "score_principal": primary_score,
                "score_pondere": weighted_score,
                "score_queue_haute": tail_score,
                "score_median": float(pd.to_numeric(group_df["score"], errors="coerce").median()),
                "n_vaisseaux": int(len(group_df)),
                "longueur_totale": float(pd.to_numeric(group_df["longueur"], errors="coerce").sum()),
            }
        )
    return pd.DataFrame(rows)


def _rank_pivot(rank_scores: pd.DataFrame, index_columns: list[str]) -> pd.DataFrame:
    if rank_scores.empty:
        return pd.DataFrame()
    pivot = rank_scores.pivot_table(index=index_columns, columns="rang", values="score_principal", aggfunc="first").reset_index()
    for rank in [1, 2, 3]:
        if rank not in pivot.columns:
            pivot[rank] = math.nan
    pivot = pivot.rename(columns={1: "score_rang_1", 2: "score_rang_2", 3: "score_rang_3"})
    counts = rank_scores.pivot_table(index=index_columns, columns="rang", values="n_vaisseaux", aggfunc="first").reset_index()
    counts = counts.rename(columns={1: "n_rang_1", 2: "n_rang_2", 3: "n_rang_3"})
    for rank in [1, 2, 3]:
        column = f"n_rang_{rank}"
        if column not in counts.columns:
            counts[column] = 0
    return pivot.merge(counts[index_columns + ["n_rang_1", "n_rang_2", "n_rang_3"]], on=index_columns, how="left")


def _patient_max_rank_scores(rank_scores: pd.DataFrame) -> pd.DataFrame:
    patient_scores = rank_scores[rank_scores["groupe"] == "patient"].copy() if not rank_scores.empty else pd.DataFrame()
    if patient_scores.empty:
        return pd.DataFrame(columns=["patient_id", "score_max_rangs_1_2_3"])
    rows = []
    for patient_id, subset in patient_scores.groupby("patient_id"):
        scores = _finite_series(subset["score_principal"])
        rows.append({"patient_id": patient_id, "score_max_rangs_1_2_3": float(scores.max()) if not scores.empty else math.nan})
    return pd.DataFrame(rows)


def _append_cohort_quality_rows(
    quality_rows: list[dict[str, object]],
    run_infos: list[RunInfo],
    seen_patient_eyes: set[tuple[int, str]],
    seen_control_eyes: set[tuple[int, str]],
    clinical: pd.DataFrame,
) -> None:
    patient_ids = sorted({info.patient_id for info in run_infos if info.group == "patient" and info.patient_id is not None})
    clinical_ids = set(clinical["patient_id"].dropna().astype(int).tolist()) if not clinical.empty and "patient_id" in clinical else set()
    for patient_id in patient_ids:
        for eye in ["OD", "OG"]:
            if (patient_id, eye) not in seen_patient_eyes:
                quality_rows.append(_quality_row(RunInfo("", patient_id, eye, "patient"), "", "oeil_patient_manquant", "Run patient OD/OG manquant"))
            if (patient_id, eye) not in seen_control_eyes:
                quality_rows.append(_quality_row(RunInfo("", patient_id, eye, "temoin"), "", "temoin_manquant", "Run temoin OD_de_x/OG_de_x manquant"))
        if patient_id not in clinical_ids:
            quality_rows.append(_quality_row(RunInfo("", patient_id, None, "patient"), "", "donnees_cliniques_manquantes", "Age/severite absents de demo/clinical_metadata.csv"))


def _append_resolution_quality_rows(quality_rows: list[dict[str, object]], data: pd.DataFrame) -> None:
    if data.empty or "diametre_fond_oeil_px" not in data:
        return
    runs = data[["run", "groupe", "patient_id", "oeil", "diametre_fond_oeil_px", "facteur_normalisation"]].drop_duplicates("run")
    for _, row in runs.iterrows():
        diameter = _finite_float(row["diametre_fond_oeil_px"])
        info = RunInfo(str(row["run"]), int(row["patient_id"]) if pd.notna(row["patient_id"]) else None, row["oeil"], str(row["groupe"]))
        if not math.isfinite(diameter):
            quality_rows.append(_quality_row(info, "", "normalisation_resolution_manquante", "Masque du fond d'oeil absent; echelle brute utilisee"))
    medians = runs.groupby("groupe")["diametre_fond_oeil_px"].median()
    if "patient" in medians and "temoin" in medians and min(medians["patient"], medians["temoin"]) > 0:
        ratio = max(medians["patient"], medians["temoin"]) / min(medians["patient"], medians["temoin"])
        if ratio >= 1.5:
            quality_rows.append(
                _quality_row(
                    RunInfo("", None, None, "cohorte"),
                    "",
                    "resolution_confondue_avec_groupe",
                    (
                        f"Diametre median brut patient={medians['patient']:.0f}px, temoin={medians['temoin']:.0f}px "
                        f"(ratio {ratio:.2f}); scores normalises sur un diametre de reference de 1024px"
                    ),
                )
            )


def _read_clinical_metadata(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=["patient_id", "age", "severite"])
    try:
        table = pd.read_csv(path)
    except (OSError, ValueError):
        return pd.DataFrame(columns=["patient_id", "age", "severite"])
    for column in ["patient_id", "age", "severite"]:
        if column not in table.columns:
            table[column] = math.nan
    table = table[["patient_id", "age", "severite"]].copy()
    table["patient_id"] = pd.to_numeric(table["patient_id"], errors="coerce").astype("Int64")
    table["age"] = pd.to_numeric(table["age"], errors="coerce")
    table["severite"] = pd.to_numeric(table["severite"], errors="coerce")
    return table.dropna(subset=["patient_id"]).drop_duplicates("patient_id", keep="first")


def _clinical_values(clinical_by_patient: pd.DataFrame, patient_id: int | None) -> dict[str, float]:
    if patient_id is None or clinical_by_patient.empty or patient_id not in clinical_by_patient.index:
        return {"age": math.nan, "severite": math.nan}
    row = clinical_by_patient.loc[patient_id]
    return {"age": _finite_float(row.get("age")), "severite": _finite_float(row.get("severite"))}


def _quality_row(run_info: RunInfo, vessel_name: str, problem: str, detail: str) -> dict[str, object]:
    return {
        "patient_id": run_info.patient_id,
        "groupe": run_info.group,
        "oeil": run_info.eye,
        "run": run_info.run,
        "vessel_name": vessel_name,
        "probleme": problem,
        "detail": detail,
    }


def _paired_summary(
    left: pd.Series,
    right: pd.Series,
    label: str,
    alternative: str = "two-sided",
) -> dict[str, object]:
    pairs = pd.DataFrame({"left": pd.to_numeric(left, errors="coerce"), "right": pd.to_numeric(right, errors="coerce")}).dropna()
    differences = pairs["left"] - pairs["right"]
    if len(pairs) >= 3 and differences.abs().sum() > 0:
        pvalue = float(wilcoxon(pairs["left"], pairs["right"], alternative="two-sided").pvalue)
        directional_pvalue = (
            float(wilcoxon(pairs["left"], pairs["right"], alternative=alternative).pvalue)
            if alternative != "two-sided"
            else math.nan
        )
    else:
        pvalue = math.nan
        directional_pvalue = math.nan
    delta = float(differences.mean()) if not pairs.empty else math.nan
    ci_low, ci_high = _bootstrap_mean_ci(differences)
    return {
        "section": "resume",
        "comparaison": label,
        "n_paires": int(len(pairs)),
        "delta_moyen": delta,
        "delta_median": float(differences.median()) if not differences.empty else math.nan,
        "proportion_delta_positif": float((differences > 0).mean()) if not differences.empty else math.nan,
        "ic95_delta_moyen_bas": ci_low,
        "ic95_delta_moyen_haut": ci_high,
        "taille_effet_biserielle": _matched_rank_biserial(differences),
        "p_value_wilcoxon": pvalue,
        "hypothese_directionnelle": alternative if alternative != "two-sided" else "",
        "p_value_directionnelle": directional_pvalue,
        "conclusion": _comparison_conclusion(pvalue, delta),
    }


def _mannwhitney_summary(left: pd.Series, right: pd.Series, label: str) -> dict[str, object]:
    left_values = _finite_series(left)
    right_values = _finite_series(right)
    if len(left_values) >= 2 and len(right_values) >= 2:
        pvalue = float(mannwhitneyu(left_values, right_values, alternative="two-sided").pvalue)
    else:
        pvalue = math.nan
    delta = float(left_values.mean() - right_values.mean()) if not left_values.empty and not right_values.empty else math.nan
    return {
        "section": "resume",
        "comparaison": label,
        "n_arteres": int(len(left_values)),
        "n_veines": int(len(right_values)),
        "delta_moyen": delta,
        "p_value_mannwhitney": pvalue,
        "conclusion": _comparison_conclusion(pvalue, delta),
    }


def _append_summary_rows(table: pd.DataFrame, summary_rows: list[dict[str, object]]) -> pd.DataFrame:
    summary = pd.DataFrame(summary_rows)
    if table.empty:
        return summary
    table = table.copy()
    table["section"] = "donnees"
    return pd.concat([summary, table], ignore_index=True, sort=False)


def _combine_sections(summary_rows: list[dict[str, object]], *tables: pd.DataFrame) -> pd.DataFrame:
    frames = [pd.DataFrame(summary_rows)]
    frames.extend(table for table in tables if not table.empty)
    return pd.concat(frames, ignore_index=True, sort=False)


def _with_summary(table: pd.DataFrame, conclusion: str) -> pd.DataFrame:
    summary = pd.DataFrame([{"section": "resume", "conclusion": conclusion}])
    if table.empty:
        return summary
    return pd.concat([table, summary], ignore_index=True, sort=False)


def _rank3_is_highest(r1: object, r2: object, r3: object) -> object:
    values = [_finite_float(r1), _finite_float(r2), _finite_float(r3)]
    if not math.isfinite(values[2]) or not (math.isfinite(values[0]) or math.isfinite(values[1])):
        return pd.NA
    return bool(values[2] >= max(value for value in values[:2] if math.isfinite(value)))


def _weighted_group_score(group: pd.DataFrame) -> float:
    if group.empty:
        return math.nan
    scores = pd.to_numeric(group["score"], errors="coerce")
    lengths = pd.to_numeric(group["longueur"], errors="coerce")
    valid = scores.notna() & lengths.notna() & (lengths > 0)
    if not valid.any():
        return math.nan
    return float((scores[valid] * lengths[valid]).sum() / lengths[valid].sum())


def _primary_group_score(group: pd.DataFrame) -> float:
    weighted_score = _weighted_group_score(group)
    if group.empty or not math.isfinite(weighted_score):
        return math.nan
    method_id = str(group["methode_id"].iloc[0]) if "methode_id" in group else ""
    if method_id != "local_bump":
        return weighted_score
    tail_score = _tail_group_score(group)
    if not math.isfinite(tail_score):
        return weighted_score
    global_weight = _finite_float(group["poids_global"].iloc[0]) if "poids_global" in group else 0.7
    tail_weight = _finite_float(group["poids_queue_haute"].iloc[0]) if "poids_queue_haute" in group else 0.3
    return global_weight * weighted_score + tail_weight * tail_score


def _tail_group_score(group: pd.DataFrame) -> float:
    if group.empty:
        return math.nan
    clean = group.assign(
        _score=pd.to_numeric(group["score"], errors="coerce"),
        _length=pd.to_numeric(group["longueur"], errors="coerce"),
    ).dropna(subset=["_score", "_length"])
    clean = clean[clean["_length"] > 0].sort_values("_score", ascending=False)
    if clean.empty:
        return math.nan
    fraction = _finite_float(clean["fraction_queue_haute"].iloc[0]) if "fraction_queue_haute" in clean else 0.2
    target_length = float(clean["_length"].sum()) * fraction
    remaining = target_length
    weighted_sum = 0.0
    used_length = 0.0
    for _, row in clean.iterrows():
        take = min(float(row["_length"]), remaining)
        weighted_sum += float(row["_score"]) * take
        used_length += take
        remaining -= take
        if remaining <= 1e-9:
            break
    return weighted_sum / used_length if used_length > 0 else math.nan


def _finite_series(values: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(values, errors="coerce").dropna()
    return numeric[numeric.map(math.isfinite)]


def _finite_float(value: object) -> float:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return math.nan
    return numeric if math.isfinite(numeric) else math.nan


def _comparison_conclusion(pvalue: float, delta: float) -> str:
    if not math.isfinite(pvalue):
        return "donnees insuffisantes"
    if pvalue < 0.05 and delta > 0:
        return "compatible avec un score plus eleve dans le premier groupe"
    if pvalue < 0.05 and delta < 0:
        return "compatible avec un score plus eleve dans le second groupe"
    return "pas de preuve statistique de difference"


def _omnibus_conclusion(pvalue: float) -> str:
    if not math.isfinite(pvalue):
        return "donnees insuffisantes"
    if pvalue < 0.05:
        return "compatible avec une difference entre au moins deux rangs"
    return "pas de preuve statistique de difference globale entre les rangs"


def _matched_rank_biserial(differences: pd.Series) -> float:
    clean = _finite_series(differences)
    clean = clean[clean != 0]
    if clean.empty:
        return math.nan
    ranks = rankdata(clean.abs().to_numpy())
    positive = float(ranks[clean.to_numpy() > 0].sum())
    negative = float(ranks[clean.to_numpy() < 0].sum())
    total = positive + negative
    return (positive - negative) / total if total > 0 else math.nan


def _bootstrap_mean_ci(values: pd.Series, repetitions: int = 5000) -> tuple[float, float]:
    clean = _finite_series(values).to_numpy(dtype=float)
    if len(clean) < 3:
        return math.nan, math.nan
    rng = np.random.default_rng(20260719)
    samples = rng.choice(clean, size=(repetitions, len(clean)), replace=True).mean(axis=1)
    low, high = np.quantile(samples, [0.025, 0.975])
    return float(low), float(high)


def _holm_adjust_summary_rows(rows: list[dict[str, object]], source: str, target: str) -> None:
    indexed = [
        (index, float(row[source]))
        for index, row in enumerate(rows)
        if source in row and math.isfinite(_finite_float(row[source]))
    ]
    if not indexed:
        return
    ordered = sorted(indexed, key=lambda item: item[1])
    running = 0.0
    adjusted: dict[int, float] = {}
    total = len(ordered)
    for order, (index, pvalue) in enumerate(ordered):
        candidate = min(1.0, pvalue * (total - order))
        running = max(running, candidate)
        adjusted[index] = running
    for index, value in adjusted.items():
        rows[index][target] = value


def _refresh_comparison_conclusions(rows: list[dict[str, object]], pvalue_column: str) -> None:
    for row in rows:
        if pvalue_column in row and "delta_moyen" in row:
            row["conclusion"] = _comparison_conclusion(
                _finite_float(row[pvalue_column]),
                _finite_float(row["delta_moyen"]),
            )


def _correlation_conclusion(pvalue: float, rho: float, variable: str) -> str:
    if not math.isfinite(pvalue):
        return "donnees insuffisantes"
    if pvalue < 0.05 and rho < 0 and variable == "age":
        return "compatible avec des scores plus eleves chez les plus jeunes"
    if pvalue < 0.05 and rho > 0:
        return "compatible avec une association positive"
    if pvalue < 0.05 and rho < 0:
        return "compatible avec une association negative"
    return "pas de preuve statistique de correlation"


def _bilateral_association_conclusion(pvalue: float, rho: float) -> str:
    if not math.isfinite(pvalue):
        return "donnees insuffisantes"
    if pvalue < 0.05 and rho > 0:
        return "association positive compatible avec une tortuosite bilaterale"
    if pvalue < 0.05 and rho < 0:
        return "association negative entre OD et OG"
    return "association OD-OG non demontree; absence de difference ne prouve pas l'equivalence"


def _normalize_name(name: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(name)).encode("ascii", "ignore").decode("ascii")
    normalized = normalized.lower().replace("°", "")
    normalized = re.sub(r"[_/]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


def _structured_columns() -> list[str]:
    return [
        "patient_id",
        "groupe",
        "oeil",
        "run",
        "vessel_name",
        "vessel_name_normalise",
        "rang",
        "type_vaisseau",
        "territoire",
        "score_brut",
        "score",
        "longueur",
        "longueur_brute_px",
        "facteur_normalisation",
        "diametre_fond_oeil_px",
        "diametre_reference_px",
        "eligible",
        "eligible_filtre_longueur",
        "nom_clinique_valide",
        "eligible_analyse",
        "raison_exclusion_analyse",
        "eligible_export_classifie",
        "raison_exclusion_export_classifie",
        "methode_id",
        "methode",
        "score_local_bump_v1",
        "score_local_bump_v2",
        "composante_oscillation_v2",
        "composante_angularite_v2",
        "nombre_lobes_persistants",
        "nombre_oscillations_persistantes",
        "tour_maximal_extremites",
        "segments_modele",
        "segments_manuels",
        "tortuosity_density_score",
        "sous_segments_courbure_constante",
        "nombre_inflexions",
        "exces_arc_corde_tortuosity_density",
        "geometrie_tortuosity_density_valide",
        "poids_global",
        "poids_queue_haute",
        "fraction_queue_haute",
        "age",
        "severite",
    ]
